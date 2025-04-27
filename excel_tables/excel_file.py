import logging
import os
import os.path
import tkinter as tk
from abc import ABC
from tkinter import filedialog
from typing import Tuple, Optional

import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from configs.utils import resource_path

log = logging.getLogger(__name__)


class ExcelFile(ABC):
    """
    Abstract base class for handling Excel files.
    
    Provides functionality for opening, manipulating, and saving Excel workbooks.
    """

    def __init__(self, file_path: str, worksheet_number: int = 0):
        """
        Initialize an Excel file handler.
        
        Parameters:
            file_path (str): Path to the Excel file to be opened
            worksheet_number (int, optional): Index of the worksheet to be used (0-based). Defaults to 0.
        """
        self.FILE_PATH = resource_path(file_path)
        self.FILE_NAME = os.path.basename(file_path)
        self.wb, self.ws = ExcelFile.open_file(self.FILE_PATH, worksheet_number)

    @staticmethod
    def open_file(file_name: str, worksheet_number: int = 0) -> Tuple[Optional[Workbook], Optional[Worksheet]]:
        """
        Open an Excel file and return the workbook and worksheet objects.
        
        Parameters:
            file_name (str): Path to the Excel file to be opened
            worksheet_number (int, optional): Index of the worksheet to be used (0-based). Defaults to 0.
            
        Returns:
            Tuple[Optional[Workbook], Optional[Worksheet]]: A tuple containing the workbook and worksheet objects,
            or (None, None) if the file could not be opened.
        """
        log.info("Opening file: %s", file_name)
        if file_name:
            try:
                wb = xl.load_workbook(file_name)
                ws = wb.worksheets[worksheet_number] if worksheet_number < len(wb.worksheets) else None
                log.info("File opened successfully: %s", file_name)
                return wb, ws
            except Exception as e:
                log.error("Error opening file %s: %s", file_name, e)
                return None, None
        else:
            return None, None

    def save_file(self, new_file_name: str = None) -> Optional[str]:
        """
        Save the workbook to a file with a dialog prompt if no filename is provided.
        
        Parameters:
            new_file_name (str, optional): Name for the saved file. If None, a file dialog will be shown.
            
        Returns:
            Optional[str]: The path where the file was saved, or None if the save operation was cancelled or failed.
            
        Raises:
            ValueError: If the workbook is not initialized.
        """
        log.info("Saving file: %s", self.FILE_NAME)
        if self.wb is None:
            raise ValueError("Workbook is not initialized")

        root = tk.Tk()
        root.withdraw()
        new_file_name = filedialog.asksaveasfilename(
            initialdir=".",
            initialfile=new_file_name,
            defaultextension=".xlsx",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        root.destroy()

        if new_file_name:
            try:
                self.wb.save(new_file_name)
                log.info("File saved successfully: %s", new_file_name)
                return new_file_name
            except Exception as e:
                log.error("Error saving file %s: %s", new_file_name, e)
        else:
            log.warning("File name not provided")
            return None

    def _write_cell(self, col: int, row: int, value) -> None:
        """
        Write a value to a specific cell in the worksheet.
        
        Parameters:
            col (int): Column number (1-based)
            row (int): Row number (1-based)
            value: The value to write to the cell
        """
        self.ws[f"{get_column_letter(col)}{row}"] = value

    def _get_cell_value(self, col: int, row: int) -> str:
        """
        Get the value from a specific cell in the worksheet.
        
        Parameters:
            col (int): Column number (1-based)
            row (int): Row number (1-based)
            
        Returns:
            str: The value of the specified cell
        """
        return self.ws[f"{get_column_letter(col)}{row}"].value

    def _group_rows(self, start_row: int, end_row: int, hidden: bool = False) -> None:
        """
        Group a range of rows in the worksheet.
        
        Parameters:
            start_row (int): First row to include in the group (1-based)
            end_row (int): Last row to include in the group (1-based)
            hidden (bool, optional): Whether the grouped rows should be hidden. Defaults to False.
        """
        self.ws.row_dimensions.group(start_row, end_row, hidden=hidden)

    def _group_columns(self, start_col: int, end_col: int, hidden: bool = False) -> None:
        """
        Group a range of columns in the worksheet.
        
        Parameters:
            start_col (int): First column to include in the group (1-based)
            end_col (int): Last column to include in the group (1-based)
            hidden (bool, optional): Whether the grouped columns should be hidden. Defaults to False.
        """
        if start_col <= end_col:
            self.ws.column_dimensions.group(get_column_letter(start_col), get_column_letter(end_col), hidden=hidden)

    def _bold_cell(self, col: int, row: int) -> None:
        """
        Apply bold formatting to a specific cell in the worksheet.
        
        Parameters:
            col (int): Column number (1-based)
            row (int): Row number (1-based)
        """
        self.ws[f"{get_column_letter(col)}{row}"].font = xl.styles.Font(bold=True, name="Calibri")
