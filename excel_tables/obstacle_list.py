import logging
from typing import Optional, Tuple, List

from openpyxl.utils import get_column_letter

from GoogleMyMaps.models import *
from configs.utils import unify_string
from .areas import Areas
from .course_trail import CourseTrail
from .courses import Courses
from .excel_file import ExcelFile

log = logging.getLogger(__name__)


class ObstacleList(ExcelFile):
    """
    A class for generating and managing obstacle lists in Excel format.
    
    This class handles the creation of an Excel file containing information about obstacles
    across different courses, including their positions, areas, distances, and required personnel.
    
    Attributes:
        IMPORTANT_OBSTACLE_NAMES (List[str]): Names of obstacles that should be highlighted in the output.
        COLUMN_LAST_COURSE (int): Column index for the last course.
        COLUMN_NAME (int): Column index for obstacle names.
        COLUMN_WOLO (int): Column index for volunteer information.
        COLUMN_JUDGE (int): Column index for judge information.
        COLUMN_INFO (int): Column index for additional obstacle information.
        ROW_HEADERS (int): Row index for headers.
        ROW_OBSTACLES_OFFSET (int): Offset for obstacle rows.
        ROW_MAX (int): Maximum row index.
        file_name (str): Base name of the Excel file.
        file_path (str): Path to the template Excel file.
        not_found_obstacles (List[Tuple[Layer, int, Place]]): List to store obstacles that couldn't be found.
    """
    
    IMPORTANT_OBSTACLE_NAMES = ["START", "META", "START KIDS", "META KIDS"]
    COLUMN_LAST_COURSE = 18
    COLUMN_NAME = 19
    COLUMN_WOLO = 20
    COLUMN_JUDGE = 21
    # COLUMN_WORKER = 22
    COLUMN_INFO = 24
    ROW_HEADERS = 1
    ROW_OBSTACLES_OFFSET = 2
    ROW_MAX = 200

    file_name = "LISTA PRZESZKÓD"
    file_path = f"WZORY/{file_name}.xlsx"

    # List to store obstacles that couldn't be found
    not_found_obstacles: List[Tuple[Layer, int, Place]] = []

    def __init__(self, google_map: Map):
        """
        Initialize the ObstacleList with a Google Map.
        
        Parameters:
            google_map (Map): The Google Map object containing course and obstacle data.
        """
        super().__init__(self.file_path)
        self.google_map = google_map
        self.courses = Courses(google_map)
        self.areas = Areas(google_map)

    def _write_headlines(self):
        """
        Write course headlines to the Excel file.
        
        Writes the shortened course names as column headers in the Excel file.
        """
        for course in self.courses.courses_list:
            self._write_cell(self._get_course_column_number(course), self.ROW_HEADERS, course.name[6:])

    def _get_course_column_number(self, course: Layer) -> int:
        """
        Calculate the column number for a given course.
        
        Parameters:
            course (Layer): The course layer to calculate the column number for.
            
        Returns:
            int: The column number where the course data should be written.
        """
        return self.courses.get_course_index(course) * 3 + 1

    def _write_obstacle_number_area_km(self, course_column: int, obstacle_row: int, obstacle_number: int,
                                       obstacle_area: int, obstacle_distance: Optional[float]) -> None:
        """
        Write obstacle number, area, and distance information to the Excel file.
        
        Parameters:
            course_column (int): The starting column number for the course.
            obstacle_row (int): The row number for the obstacle.
            obstacle_number (int): The obstacle's number.
            obstacle_area (int): The area number where the obstacle is located.
            obstacle_distance (Optional[float]): The distance to the obstacle in meters, or None if unknown.
        """
        if obstacle_distance is not None:
            obstacle_distance = round(obstacle_distance / 1000, 1)
        self._write_cell(course_column, obstacle_row, obstacle_number)
        self._write_cell(course_column + 1, obstacle_row, obstacle_area)
        self._write_cell(course_column + 2, obstacle_row, obstacle_distance)

    def _write_obstacle_name_data(self, obstacle: Place, obstacle_row: int):
        """
        Write obstacle name and associated data to the Excel file.
        
        Parameters:
            obstacle (Place): The obstacle place object containing name and data.
            obstacle_row (int): The row number for the obstacle.
        """
        self._write_cell(self.COLUMN_NAME, obstacle_row, obstacle.name)
        if obstacle.name.upper() in self.IMPORTANT_OBSTACLE_NAMES:
            self._bold_cell(self.COLUMN_NAME, obstacle_row)

        # self.ws["V" + str(cell_line)] = # Responsible person

        data = self._get_obstacle_data(obstacle)
        if data is None:
            return

        wolo, judge, description = data
        self._write_cell(self.COLUMN_WOLO, obstacle_row, wolo if wolo > 0 else None)
        self._write_cell(self.COLUMN_JUDGE, obstacle_row, judge if judge > 0 else None)
        self._write_cell(self.COLUMN_INFO, obstacle_row, description)

    def _get_obstacle_data(self, obstacle: Place) -> Optional[Tuple[int, int, str]]:
        """
        Extract volunteer, judge, and description data from an obstacle.
        
        Parameters:
            obstacle (Place): The obstacle place object to extract data from.
            
        Returns:
            Optional[Tuple[int, int, str]]: A tuple containing (volunteer count, judge count, description),
                                           or None if the obstacle has no data.
        """
        if obstacle.data is None:
            return None
        normalized_data = {key[0].upper(): value for key, value in obstacle.data.items()}

        wolo = self._get_person_number(normalized_data, "W")  # WOLO
        judge = self._get_person_number(normalized_data, "S")  # SĘDZIA
        description = normalized_data.get("O", "")  # OPIS
        return wolo, judge, description

    @staticmethod
    def _get_person_number(data: dict, data_key: str) -> int:
        """
        Extract and convert a person count from obstacle data.
        
        Parameters:
            data (dict): The normalized obstacle data dictionary.
            data_key (str): The key to look up in the data dictionary.
            
        Returns:
            int: The number of persons, or 0 if not found or invalid.
        """
        try:
            return int(data.get(data_key, 0))
        except ValueError:
            # TODO: info?
            log.warning("Invalid data type for %s: %s", data_key, type(data.get(data_key, 0)))
            return 0

    def _write_course_info(self, course: Layer):
        """
        Write all obstacle information for a specific course.
        
        Parameters:
            course (Layer): The course layer containing obstacles to write.
        """
        row_offset = self.courses.get_course_obstacles_number(self.courses.courses_list[
                                                                  0]) + self.ROW_OBSTACLES_OFFSET + 2 if "KIDS" in course.name.upper() else self.ROW_OBSTACLES_OFFSET

        for obstacle in course.places:
            if obstacle.place_type != "Point":
                continue

            self._write_single_obstacle_info(course, obstacle, row_offset)

    def _write_single_obstacle_info(self, course: Layer, obstacle: Place, row_offset: int):
        """
        Write information for a single obstacle in a course.
        
        Parameters:
            course (Layer): The course layer the obstacle belongs to.
            obstacle (Place): The obstacle place object to write information for.
            row_offset (int): The row offset to apply when calculating the obstacle's row.
        """
        course_column = self._get_course_column_number(course)
        obstacle_number = self.courses.get_obstacle_number(obstacle)
        if obstacle_number is None:
            log.warning("Obstacle without number: %s, %s", obstacle.name, obstacle.icon)
            return
        obstacle_row = obstacle_number + row_offset
        obstacle_area_number = self.areas.get_obstacle_area_number(obstacle)
        obstacle_distance = CourseTrail(course).get_obstacle_distance(obstacle)
        self._write_obstacle_number_area_km(course_column, obstacle_row, obstacle_number, obstacle_area_number, obstacle_distance)
        self._write_obstacle_name_data(obstacle, obstacle_row)

    def _write_obstacles_numbers(self, course: Layer):
        """
        Write obstacle numbers for a course by matching obstacles with the main courses.
        
        Parameters:
            course (Layer): The course layer to process obstacles for.
        """
        obstacle_row_offset = self.ROW_OBSTACLES_OFFSET
        kids_obstacle_row_offset = self.courses.get_course_obstacles_number(
            self.courses.courses_list[0]) + self.ROW_OBSTACLES_OFFSET + 2

        last_found_obstacle_index = -1
        for analysed_obstacle in course.places:
            if analysed_obstacle.place_type != "Point":
                continue

            last_found_obstacle_index = self._process_obstacle(
                course,
                analysed_obstacle,
                obstacle_row_offset,
                kids_obstacle_row_offset,
                last_found_obstacle_index
            )

    def _process_obstacle(
            self,
            course: Layer,
            analysed_obstacle: Place,
            obstacle_row_offset: int,
            kids_obstacle_row_offset: int,
            last_found_obstacle_index: int
    ) -> int:
        """
        Process an obstacle by trying to find a matching obstacle in the main courses.
        
        Parameters:
            course (Layer): The course layer the obstacle belongs to.
            analysed_obstacle (Place): The obstacle place object to process.
            obstacle_row_offset (int): The row offset for the main course.
            kids_obstacle_row_offset (int): The row offset for the kids course.
            last_found_obstacle_index (int): The index of the last found obstacle.
            
        Returns:
            int: The updated index of the last found obstacle.
        """
        found_obstacle_index = self._find_and_write_obstacle(
            analysed_obstacle,
            self.courses.courses_list[0].places[last_found_obstacle_index + 1:],
            course,
            obstacle_row_offset,
        )
        if found_obstacle_index is not None:
            return last_found_obstacle_index + found_obstacle_index + 1

        found_obstacle_index = self._find_and_write_obstacle(
            analysed_obstacle,
            self.courses.courses_list[0].places[last_found_obstacle_index + 1::-1],
            course,
            obstacle_row_offset,
        )
        if found_obstacle_index is not None:
            return last_found_obstacle_index + found_obstacle_index + 1

        found_obstacle_index = self._find_and_write_obstacle(
            analysed_obstacle,
            self.courses.courses_list[-1].places,
            course,
            kids_obstacle_row_offset,
        )
        if found_obstacle_index is None:
            log.warning("-Unable to find obstacle: %s from course: %s", analysed_obstacle.name, course.name)
            # Add to list with course name
            self.add_to_list(analysed_obstacle, course)

        return last_found_obstacle_index

    def _find_and_write_obstacle(self, analysed_obstacle: Place, obstacles, course: Layer, row_offset: int) -> \
            Optional[int]:
        """
        Find a matching obstacle in a list and write its information.
        
        Parameters:
            analysed_obstacle (Place): The obstacle place object to find a match for.
            obstacles: The list of obstacles to search in.
            course (Layer): The course layer the analysed obstacle belongs to.
            row_offset (int): The row offset to apply when calculating the obstacle's row.
            
        Returns:
            Optional[int]: The index of the found obstacle, or None if not found.
        """
        for obstacle in obstacles:
            if unify_string(analysed_obstacle.name) == unify_string(obstacle.name):
                obstacle_number = self.courses.get_obstacle_number(obstacle)
                if obstacle_number is None:
                    log.warning("Obstacle without number: %s, %s", obstacle.name, obstacle.icon)
                    return
                obstacle_row = obstacle_number + row_offset
                course_column = self._get_course_column_number(course)
                if self._get_cell_value(course_column, obstacle_row) is not None:
                    continue

                analysed_obstacle_number = self.courses.get_obstacle_number(analysed_obstacle)
                obstacle_area_number = self.areas.get_obstacle_area_number(obstacle)
                obstacle_distance = CourseTrail(course).get_obstacle_distance(obstacle)
                self._write_obstacle_number_area_km(course_column, obstacle_row, analysed_obstacle_number,
                                                    obstacle_area_number, obstacle_distance)

                return obstacles.index(obstacle)
        return None

    def _sum_and_write_number_of_volunteers_and_judges(self):
        """
        Calculate and write the total number of volunteers and judges.
        
        Adds formulas to sum the volunteer and judge columns in the Excel file.
        """
        self._write_cell(
            self.COLUMN_WOLO,
            self.ROW_MAX + 1,
            f"=SUM({get_column_letter(self.COLUMN_WOLO)}{self.ROW_OBSTACLES_OFFSET + 1}:{get_column_letter(self.COLUMN_WOLO)}{self.ROW_MAX})"
        )
        self._write_cell(
            self.COLUMN_JUDGE,
            self.ROW_MAX + 1,
            f"=SUM({get_column_letter(self.COLUMN_JUDGE)}{self.ROW_OBSTACLES_OFFSET + 1}:{get_column_letter(self.COLUMN_JUDGE)}{self.ROW_MAX})"
        )

    def _hide_unnecessary_columns_and_rows(self):
        """
        Hide unnecessary columns and rows in the Excel file.
        
        Groups and hides columns and rows that don't contain relevant data to improve readability.
        """
        self._group_columns(
            self._get_course_column_number(self.courses.courses_list[-1]) + 3,
            self.COLUMN_LAST_COURSE,
            True
        )
        self._group_rows(
            self.courses.get_course_obstacles_number(self.courses.courses_list[0])
            + self.courses.get_course_obstacles_number(self.courses.courses_list[-1])
            + self.ROW_OBSTACLES_OFFSET + 2,
            self.ROW_MAX,
            True
        )

    def _save_data(self) -> Optional[str]:
        """
        Save all obstacle data to the Excel file.
        
        This method orchestrates the process of writing all obstacle data to the Excel file
        and saving it with an appropriate name.
        
        Returns:
            Optional[str]: The path to the saved file, or None if saving failed.
        """
        self._write_headlines()
        self._write_course_info(self.courses.courses_list[0])
        self._write_course_info(self.courses.courses_list[-1])
        for course in self.courses.courses_list[1:-1]:
            self._write_obstacles_numbers(course)
        self._sum_and_write_number_of_volunteers_and_judges()
        self._hide_unnecessary_columns_and_rows()
        return self.save_file(self.google_map.name + " - LISTA PRZESZKÓD.xlsx")

    def add_to_list(self, obstacle: Place, course: Layer):
        """
        Add an obstacle to the list of obstacles that couldn't be found in the main courses.
        
        This method tracks obstacles that couldn't be matched with obstacles in the main courses,
        storing them for later reporting or analysis.
        
        Parameters:
            obstacle (Place): The obstacle place object that couldn't be found.
            course (Layer): The course layer the obstacle belongs to.
        
        Returns:
            None: This method modifies the internal not_found_obstacles list but doesn't return a value.
        """
        self.not_found_obstacles.append((course, self.courses.get_obstacle_number(obstacle), obstacle))

    def create_and_save(self) -> Tuple[Optional[str], List[Tuple[Layer, int, Place]]]:
        """
        Creates and saves the obstacle list Excel file.
        
        This method orchestrates the entire process of generating the obstacle list Excel file:
        1. Writes the course headlines
        2. Writes detailed information for the main and kids courses
        3. Processes all intermediate courses
        4. Calculates and writes totals for volunteers and judges
        5. Hides unnecessary columns and rows for better readability
        6. Saves the file with the map name
        
        Returns:
            Tuple[Optional[str], List[Tuple[Layer, int, Place]]]: A tuple containing:
                - The path to the saved Excel file, or None if saving failed
                - A list of obstacles that couldn't be found, each represented as a tuple of
                  (course layer, obstacle number, obstacle place)
        """
        file_path = self._save_data()
        return file_path, self.not_found_obstacles
